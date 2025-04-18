import os
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from astral import LocationInfo
from astral.sun import sun
import pytz

# 📅 오늘 기준 다음 토요일부터 시작하는 날짜 계산 함수
def get_saturdays(start_date, count=13):
    days_until_saturday = (5 - start_date.weekday()) % 7
    first_saturday = start_date + timedelta(days=days_until_saturday)
    return [first_saturday + timedelta(weeks=i) for i in range(count)]

# ✅ 분기 계산 함수
def get_quarter(month: int) -> int:
    return ((month - 1) // 3) + 1

# 🗂️ 폴더 생성
output_dir = "excel_result"
os.makedirs(output_dir, exist_ok=True)

# 📁 JSON 불러오기 (pandas로 처리)
with open("data.json", "r", encoding="utf-8") as f:
    json_data = json.load(f)

# 📆 기준 날짜: 기준일을 문자열로 지정
base_date_str = "2025-04-01"
today = datetime.strptime(base_date_str, "%Y-%m-%d")
saturdays = get_saturdays(today)

# 📍 치바시 기준 일몰시간 계산 함수
def get_sunset(dt: datetime):
    city = LocationInfo(
        name="Chiba",
        region="Japan",
        timezone="Asia/Tokyo",
        latitude=35.6074,
        longitude=140.1065
    )
    s = sun(city.observer, date=dt.date(), tzinfo=pytz.timezone(city.timezone))
    return s['sunset'].strftime("%H:%M")

# 📌 주차 반복 (1thWeek ~ 13thWeek)
for i in range(13):
    week_key = f"{i+1}thWeek"
    week_data = json_data[week_key]
    
    # 📄 템플릿 파일 열기
    wb = load_workbook("templates/template.xlsx")
    ws = wb["inputdata"]

    # 📥 DataFrame으로 변환 (데이터 핸들링을 쉽게)
    sabbath_df = pd.json_normalize(week_data['sabbathSchool'])
    worship_df = pd.json_normalize(week_data['worshipService'])
    schedule_df = pd.DataFrame(week_data['schedule'])

    # ✅ G2 셀에 들어갈 nakaTitle 계산
    saturday = saturdays[i]
    quarter = get_quarter(saturday.month)
    date_str = f"{saturday.year}年 {saturday.month}月 {saturday.day}日"
    nakaTitle = f"{date_str}　第 {quarter}期　{i+1}回目　安息日礼拝案内"
    ws["G2"] = nakaTitle

    # ✅ G3 셀에 일몰시간 입력
    sunset_time = get_sunset(saturday)
    ws["G3"] = f"日没/Sunset {sunset_time}"

    # 🔽 sabbathSchool 입력
    ws["C2"] = sabbath_df.at[0, "reception"]
    ws["C3"] = sabbath_df.at[0, "pianist"]
    ws["C4"] = sabbath_df.at[0, "greetings"]
    ws["C6"] = sabbath_df.at[0, "miniPrayerTime"]
    ws["C9"] = sabbath_df.at[0, "program"]

    # 🔽 worshipService 입력
    ws["C14"] = worship_df.at[0, "pianist"]
    ws["C15"] = worship_df.at[0, "translator"]
    ws["C16"] = worship_df.at[0, "presider"]
    ws["C21"] = worship_df.at[0, "specialMusic"]
    ws["C22"] = worship_df.at[0, "preacher"]
    ws["C26"] = worship_df.at[0, "offeringService"]
    ws["C27"] = worship_df.at[0, "offeringPrayer"]

    # 🔽 schedule 입력 (1st: B열, 2nd: C열, 3rd: D열)
    col_map = {"1st": "B", "2nd": "C", "3rd": "D"}
    for key, col in col_map.items():
        row = schedule_df[key]
        ws[f"{col}33"] = row.get("date")
        ws[f"{col}34"] = row.get("reception")
        ws[f"{col}35"] = row.get("pianist")
        ws[f"{col}36"] = row.get("greetings")
        ws[f"{col}37"] = row.get("program")
        ws[f"{col}38"] = row.get("announcementTranslator")
        ws[f"{col}39"] = row.get("presider")
        ws[f"{col}40"] = row.get("specialMusic")
        ws[f"{col}41"] = row.get("preacher")
        ws[f"{col}42"] = row.get("SermonTranslator")
        ws[f"{col}43"] = row.get("offeringService")
        ws[f"{col}44"] = row.get("offeringPrayer")

    # 💾 저장 경로 설정
    mmdd = saturdays[i].strftime("%m%d")
    file_name = f"bulletin_{mmdd}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # 💾 저장
    wb.save(file_path)

print("✅ 13개의 엑셀 파일이 성공적으로 생성되었습니다!")
